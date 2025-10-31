# app/api/incomes.py
from fastapi import APIRouter, HTTPException, Body
from fastapi.responses import StreamingResponse
from app.core.firebase import db
from google.cloud import firestore
from typing import Dict, Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from datetime import datetime

router = APIRouter()


@router.get("/incomes")
def list_incomes():
    """
    Lista todas as receitas — manuais e automáticas (do financeiro).
    """
    try:
        incomes_ref = db.collection("incomes").order_by("date", direction=firestore.Query.DESCENDING).stream()
        reservations_ref = db.collection("reservations").stream()

        incomes = []

        # 🔹 Receitas manuais
        for doc in incomes_ref:
            data = doc.to_dict() or {}
            incomes.append({
                "id": doc.id,
                "description": data.get("description"),
                "date": data.get("date"),
                "amount": data.get("amount"),
                "method": data.get("method"),
                "origin": "Manual",
            })

        # 🔸 Receitas automáticas (pagamentos confirmados)
        for res in reservations_ref:
            data = res.to_dict() or {}
            status = (data.get("paymentStatus") or data.get("status") or "").lower()
            if any(k in status for k in ["confirmado", "pago", "aprovado"]):
                amount = float(data.get("amountReceived") or data.get("value") or 0)
                if amount > 0:
                    incomes.append({
                        "id": res.id,
                        "description": f"Reserva - {data.get('guestName') or data.get('companyName') or 'Cliente'}",
                        "date": data.get("checkOut"),
                        "amount": amount,
                        "method": data.get("paymentMethod") or "Outros",
                        "origin": "Automática",
                    })

        return sorted(incomes, key=lambda x: x["date"], reverse=True)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/incomes")
def create_income(payload: Dict[str, Any] = Body(...)):
    """
    Adiciona uma nova receita manual.
    """
    try:
        description = payload.get("description")
        date = payload.get("date")
        amount = float(payload.get("amount", 0))
        method = payload.get("method")

        if not all([description, date, amount, method]):
            raise HTTPException(status_code=400, detail="Campos obrigatórios ausentes.")

        db.collection("incomes").add({
            "description": description,
            "date": date,
            "amount": amount,
            "method": method,
            "createdAt": firestore.SERVER_TIMESTAMP,
        })

        return {"message": "Receita adicionada com sucesso."}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/incomes/export")
def export_incomes():
    """
    Exporta todas as receitas (manuais e automáticas) em planilha Excel (.xlsx)
    com cabeçalhos e formatação em português.
    """
    try:
        incomes = list_incomes()

        wb = Workbook()
        ws = wb.active
        ws.title = "Receitas"

        # Cabeçalhos
        headers = ["ID", "Descrição", "Data", "Valor (R$)", "Método de Pagamento", "Origem"]
        ws.append(headers)

        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for col_num, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.style = header_style

        # Dados
        for inc in incomes:
            valor_formatado = f"R$ {float(inc['amount']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            data_formatada = (
                datetime.strptime(inc["date"], "%Y-%m-%d").strftime("%d/%m/%Y")
                if inc.get("date") else ""
            )

            ws.append([
                inc["id"],
                inc["description"],
                data_formatada,
                valor_formatado,
                inc["method"],
                inc["origin"],
            ])

        # Ajustar largura automática
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Borda
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # Exportar para memória
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="receitas.xlsx"'
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
